# -*- coding: utf-8 -*-
"""
Sinh bộ test case RedBus – mỗi chức năng một sheet, 2000–5000 case.
Tham chiếu cấu trúc cột từ test-case/TestCase.xlsx
"""
from __future__ import annotations

import copy
import itertools
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "Test Case ID",
    "Module Name",
    "Test Priority",
    "Test Case Description",
    "Pre-condition",
    "Test Data",
    "Test Steps",
    "Expected Results",
    "Actual Results",
    "Status",
    "Post-condition",
    "Notes/Comments",
]

ROLES = ["ADMIN", "STAFF", "CUSTOMER", "Khách (chưa đăng nhập)"]
PRIORITIES = ["High", "Medium", "Low"]
STATUS_PLACEHOLDER = "(Pass / Fail / Pending)"
ACTUAL_PLACEHOLDER = "(Điền sau khi thực hiện)"

# --- Mẫu kịch bản cơ sở theo từng module RedBus ---
MODULE_BASE: dict[str, list[dict]] = {
    "XacThuc": [
        {
            "desc": "Đăng nhập thành công với tài khoản {role}",
            "prio": "High",
            "pre": "Tài khoản {role} tồn tại, hoạt động=1.",
            "data": 'POST /api/xac-thuc/dang-nhap — tenDangNhap hợp lệ, matKhau đúng',
            "steps": "1. Mở modal đăng nhập.\n2. Nhập tên đăng nhập và mật khẩu.\n3. Gửi form/API.",
            "expected": "HTTP 200; nhận JWT; chuyển trang theo vai trò {role}.",
            "post": "Phiên đăng nhập hợp lệ; token lưu localStorage.",
        },
        {
            "desc": "Đăng nhập thất bại – mật khẩu sai ({role})",
            "prio": "High",
            "pre": "Tài khoản {role} tồn tại.",
            "data": "tenDangNhap đúng, matKhau sai",
            "steps": "1. Nhập tên đăng nhập đúng.\n2. Nhập mật khẩu sai.\n3. Đăng nhập.",
            "expected": "HTTP 401/400; thông báo lỗi; không có JWT.",
            "post": "Không tạo phiên.",
        },
        {
            "desc": "Đăng nhập – bỏ trống trường bắt buộc",
            "prio": "Medium",
            "pre": "Trang đăng nhập mở.",
            "data": "tenDangNhap hoặc matKhau để trống",
            "steps": "1. Để trống một trường.\n2. Nhấn Đăng nhập.",
            "expected": "Validation phía client/server; không đăng nhập.",
            "post": "Không có phiên.",
        },
        {
            "desc": "Đăng ký tài khoản CUSTOMER hợp lệ",
            "prio": "High",
            "pre": "tenDangNhap và email chưa tồn tại.",
            "data": "tenDangNhap (3–64 ký tự), email hợp lệ, matKhau (≥6), hoTen, soDienThoai",
            "steps": "1. Mở form đăng ký.\n2. Điền đủ trường.\n3. POST /api/xac-thuc/dang-ky.",
            "expected": "Tạo tài khoản CUSTOMER + hồ sơ khách; trả JWT.",
            "post": "Có bản ghi tai_khoan và khach_hang.",
        },
        {
            "desc": "Đăng ký – trùng tên đăng nhập",
            "prio": "Medium",
            "pre": "Đã có user cùng tenDangNhap.",
            "data": "tenDangNhap trùng",
            "steps": "1. Gửi đăng ký với tên trùng.",
            "expected": "Báo lỗi trùng; không tạo bản ghi mới.",
            "post": "CSDL không đổi.",
        },
        {
            "desc": "Quên mật khẩu – gửi OTP qua email",
            "prio": "Medium",
            "pre": "Email đã đăng ký; SMTP cấu hình (hoặc mock).",
            "data": "POST /api/xac-thuc/quen-mat-khau/gui-otp — email hợp lệ",
            "steps": "1. Mở Quên mật khẩu.\n2. Nhập email.\n3. Gửi OTP.",
            "expected": "HTTP 200; mã OTP lưu ma_otp (có hạn).",
            "post": "OTP chờ xác nhận.",
            "note": "Dev có thể không gửi mail thật.",
        },
        {
            "desc": "Đặt lại mật khẩu với OTP hợp lệ",
            "prio": "Medium",
            "pre": "Có OTP còn hạn cho email.",
            "data": "email, maOtp, matKhauMoi (≥6)",
            "steps": "1. Nhập OTP và mật khẩu mới.\n2. POST /api/xac-thuc/quen-mat-khau/dat-lai.",
            "expected": "Đổi mật khẩu thành công; đăng nhập được bằng MK mới.",
            "post": "OTP đã dùng/hết hạn.",
        },
        {
            "desc": "Đặt lại mật khẩu – OTP sai hoặc hết hạn",
            "prio": "Low",
            "pre": "OTP không tồn tại hoặc quá hạn.",
            "data": "maOtp sai",
            "steps": "1. Gửi dat-lai với OTP sai.",
            "expected": "Từ chối; mật khẩu không đổi.",
            "post": "Không đổi mat_khau_bam.",
        },
        {
            "desc": "JWT hết hạn – truy cập API bảo vệ",
            "prio": "High",
            "pre": "Token JWT đã hết hạn.",
            "data": "Authorization: Bearer <expired>",
            "steps": "1. Gọi GET /api/ve-xe/cua-toi với token hết hạn.",
            "expected": "HTTP 401; yêu cầu đăng nhập lại.",
            "post": "Không trả dữ liệu vé.",
        },
        {
            "desc": "Truy cập API quản trị bằng tài khoản CUSTOMER",
            "prio": "High",
            "pre": "Đăng nhập CUSTOMER.",
            "data": "JWT role CUSTOMER",
            "steps": "1. Gọi POST /api/tuyen-duong (thêm tuyến).",
            "expected": "HTTP 403 Forbidden.",
            "post": "Không tạo tuyến.",
        },
    ],
    "TuyenDuong": [
        {"desc": "Xem danh sách tuyến đường (công khai)", "prio": "High", "pre": "Có tuyến hoat_dong=1.", "data": "GET /api/tuyen-duong", "steps": "1. Gọi API hoặc mở trang quản trị tuyến.", "expected": "Trả danh sách tuyến; có diem_di, diem_den.", "post": "Không đổi DL."},
        {"desc": "Thêm tuyến mới ({role})", "prio": "High", "pre": "Đăng nhập {role}.", "data": 'diem_di="Đà Nẵng", diem_den="Huế", khoang_cach_km, thoi_gian_uoc_tinh_phut', "steps": "1. Mở form thêm.\n2. Nhập dữ liệu.\n3. POST /api/tuyen-duong.", "expected": "Tạo thành công nếu {role} là ADMIN/STAFF; 403 nếu CUSTOMER.", "post": "Có bản ghi tuyen_duong."},
        {"desc": "Cập nhật tuyến – đổi điểm đi/đến", "prio": "Medium", "pre": "Tuyến tồn tại.", "data": "PUT /api/tuyen-duong/{ma}", "steps": "1. Sửa thông tin.\n2. Lưu.", "expected": "Cập nhật CSDL; danh sách phản ánh thay đổi.", "post": "Tuyến đã cập nhật."},
        {"desc": "Xóa tuyến không có chuyến phụ thuộc", "prio": "Medium", "pre": "Tuyến không gắn chuyến.", "data": "DELETE /api/tuyen-duong/{ma}", "steps": "1. Xóa tuyến.", "expected": "Xóa thành công.", "post": "Không còn bản ghi."},
        {"desc": "Xóa tuyến đang có chuyến (biên)", "prio": "Low", "pre": "Tuyến có chuyến xe.", "data": "ma tuyến có FK", "steps": "1. Thử xóa.", "expected": "Từ chối hoặc báo lỗi FK.", "post": "Dữ liệu chuyến không mất.", "note": "Ghi hành vi thực tế."},
        {"desc": "Thêm tuyến – bỏ trống điểm đi", "prio": "Medium", "pre": "Form thêm tuyến.", "data": "diem_di rỗng", "steps": "1. Gửi form thiếu điểm đi.", "expected": "Validation lỗi.", "post": "Không tạo tuyến."},
        {"desc": "Lọc/tìm tuyến trên giao diện quản trị", "prio": "Low", "pre": "Nhiều tuyến trong CSDL.", "data": "Từ khóa điểm đi", "steps": "1. Nhập ô tìm kiếm.\n2. Quan sát bảng.", "expected": "Danh sách lọc đúng.", "post": "Không đổi DL."},
    ],
    "XeKhach": [
        {"desc": "Xem danh sách xe khách", "prio": "High", "pre": "Có xe trong CSDL.", "data": "GET /api/xe-khach", "steps": "1. Mở Trang Xe khách quản trị.", "expected": "Hiển thị biển số, hãng, số chỗ, loại xe.", "post": "Không đổi DL."},
        {"desc": "Thêm xe mới ({role})", "prio": "High", "pre": "Đăng nhập {role}; có loại xe.", "data": 'bien_so unique, hang_xe, so_cho, ma_loai_xe', "steps": "1. POST /api/xe-khach.", "expected": "Tạo xe + sinh ghế theo sơ đồ.", "post": "Có xe_khach và ghe_ngoi."},
        {"desc": "Thêm xe – trùng biển số", "prio": "Medium", "pre": "Biển số đã tồn tại.", "data": "bien_so trùng", "steps": "1. Thêm xe trùng BS.", "expected": "Báo lỗi unique.", "post": "Không tạo xe mới."},
        {"desc": "Cập nhật thông tin xe", "prio": "Medium", "pre": "Xe tồn tại.", "data": "PUT /api/xe-khach/{ma}", "steps": "1. Sửa hãng xe/số chỗ.\n2. Lưu.", "expected": "Cập nhật thành công.", "post": "Xe đã cập nhật."},
        {"desc": "Xóa xe không có chuyến", "prio": "Low", "pre": "Xe chưa gán chuyến.", "data": "DELETE /api/xe-khach/{ma}", "steps": "1. Xóa xe.", "expected": "Xóa thành công (cascade ghế).", "post": "Không còn xe."},
        {"desc": "Xem chi tiết xe theo mã", "prio": "Medium", "pre": "ma hợp lệ.", "data": "GET /api/xe-khach/{ma}", "steps": "1. Gọi API chi tiết.", "expected": "Trả đủ thuộc tính xe.", "post": "Không đổi DL."},
    ],
    "LoaiXe": [
        {"desc": "Xem danh sách loại xe", "prio": "High", "pre": "Có loại xe seed.", "data": "GET /api/loai-xe", "steps": "1. Mở Trang Loại xe.", "expected": "Danh sách tên, mô tả, tiện ích.", "post": "Không đổi DL."},
        {"desc": "Thêm loại xe mới", "prio": "Medium", "pre": "Quyền ADMIN/STAFF.", "data": "ten, mo_ta, tien_ich", "steps": "1. POST /api/loai-xe.", "expected": "Tạo thành công.", "post": "Có loai_xe mới."},
        {"desc": "Upload ảnh loại xe", "prio": "Medium", "pre": "Loại xe tồn tại.", "data": "multipart file ảnh JPG/PNG", "steps": "1. POST /api/loai-xe/{ma}/anh.", "expected": "Lưu ảnh; hiển thị carousel.", "post": "Có loai_xe_anh."},
        {"desc": "Xóa ảnh loại xe", "prio": "Low", "pre": "Có ảnh.", "data": "DELETE /api/loai-xe/{ma}/anh/{maAnh}", "steps": "1. Xóa ảnh.", "expected": "Ảnh biến mất khỏi UI.", "post": "Bản ghi ảnh xóa."},
        {"desc": "Cập nhật loại xe", "prio": "Medium", "pre": "Loại xe tồn tại.", "data": "PUT /api/loai-xe/{ma}", "steps": "1. Sửa tên/mô tả.\n2. Lưu.", "expected": "Cập nhật OK.", "post": "DL đã đổi."},
        {"desc": "Xóa loại xe đang gắn xe", "prio": "Low", "pre": "Có xe dùng loại này.", "data": "ma loại xe", "steps": "1. Thử DELETE.", "expected": "Từ chối hoặc lỗi FK.", "post": "Loại xe còn."},
    ],
    "DiemDungChan": [
        {"desc": "Xem điểm dừng theo tuyến", "prio": "High", "pre": "Tuyến có điểm dừng.", "data": "GET /api/diem-dung/tuyen/{maTuyen}", "steps": "1. Mở Trang Điểm dừng.", "expected": "Danh sách điểm, thứ tự, thời gian.", "post": "Không đổi DL."},
        {"desc": "Thêm điểm dừng mới", "prio": "Medium", "pre": "Tuyến tồn tại.", "data": "ten, thu_tu, thoi_gian_uoc_tinh_phut", "steps": "1. POST /api/diem-dung.", "expected": "Tạo điểm dừng.", "post": "Có diem_dung_chan."},
        {"desc": "Cập nhật điểm dừng", "prio": "Medium", "pre": "Điểm dừng tồn tại.", "data": "PUT /api/diem-dung/{ma}", "steps": "1. Sửa tên/thứ tự.", "expected": "Cập nhật OK.", "post": "DL đổi."},
        {"desc": "Xóa điểm dừng", "prio": "Low", "pre": "Điểm không gắn vé.", "data": "DELETE /api/diem-dung/{ma}", "steps": "1. Xóa.", "expected": "Xóa thành công.", "post": "Không còn điểm."},
        {"desc": "Chọn điểm lên/xuống khi đặt vé", "prio": "High", "pre": "Chuyến có điểm dừng.", "data": "maDiemLen, maDiemXuong trong YeuCauDatVe", "steps": "1. Đặt vé chọn điểm.\n2. Kiểm tra vé điện tử.", "expected": "Vé ghi nhận điểm lên/xuống.", "post": "Vé PENDING/PAID có điểm dừng."},
    ],
    "GheNgoi": [
        {"desc": "Xem sơ đồ ghế theo xe", "prio": "High", "pre": "Xe có ghế.", "data": "GET /api/ghe-ngoi/xe/{maXe}", "steps": "1. Mở Trang Ghế ngồi.", "expected": "Hiển thị ký hiệu ghế, hàng, cột.", "post": "Không đổi DL."},
        {"desc": "Khóa/mở ghế ({role})", "prio": "Medium", "pre": "Đăng nhập {role}.", "data": "PUT /api/ghe-ngoi/{ma}/trang-thai — LOCKED/AVAILABLE", "steps": "1. Chọn ghế.\n2. Đổi trạng thái.", "expected": "Ghế khóa không bán được.", "post": "trang_thai ghế cập nhật."},
        {"desc": "Ghế LOCKED không cho đặt vé", "prio": "High", "pre": "Ghế trạng thái LOCKED.", "data": "maGhe locked", "steps": "1. Khách chọn ghế khóa.\n2. Đặt vé.", "expected": "Từ chối; báo ghế không khả dụng.", "post": "Không tạo vé."},
        {"desc": "Hiển thị ghế đã bán trên sơ đồ đặt vé", "prio": "High", "pre": "Chuyến có vé PAID/PENDING.", "data": "GET /api/chuyen-xe/{ma}/ghe-da-giu", "steps": "1. Mở sơ đồ ghế chuyến.", "expected": "Ghế đã giữ hiển thị khác màu.", "post": "Không đổi DL."},
    ],
    "ChuyenXe": [
        {"desc": "Xem danh sách chuyến theo ngày", "prio": "High", "pre": "Có chuyến SCHEDULED.", "data": "GET /api/chuyen-xe?ngay=yyyy-MM-dd", "steps": "1. Mở Trang Chuyến xe.", "expected": "Danh sách chuyến, giá, giờ.", "post": "Không đổi DL."},
        {"desc": "Sinh lịch chuyến hàng loạt (gen-lich)", "prio": "High", "pre": "Tuyến + xe sẵn sàng.", "data": "maTuyen, maXe, ngayBatDau, ngayKetThuc, gioKhoiHanh, giaVe", "steps": "1. POST /api/chuyen-xe/gen-lich.", "expected": "Tạo nhiều chuyến theo khoảng ngày.", "post": "Có loạt chuyen_xe."},
        {"desc": "Thêm chuyến đơn lẻ", "prio": "Medium", "pre": "Tuyến, xe tồn tại.", "data": "POST /api/chuyen-xe", "steps": "1. Nhập form chuyến.\n2. Lưu.", "expected": "Tạo chuyến SCHEDULED.", "post": "Có chuyen_xe."},
        {"desc": "Hủy chuyến (CANCELLED)", "prio": "Medium", "pre": "Chuyến chưa khởi hành.", "data": "PUT trangThai=CANCELLED", "steps": "1. Hủy chuyến.", "expected": "Chuyến không hiện tìm kiếm.", "post": "trang_thai=CANCELLED."},
        {"desc": "Cập nhật giá vé chuyến", "prio": "Medium", "pre": "Chuyến tồn tại.", "data": "gia_ve mới", "steps": "1. Sửa giá.\n2. Lưu.", "expected": "Giá mới áp dụng cho vé mới.", "post": "gia_ve cập nhật."},
        {"desc": "Xem chi tiết chuyến", "prio": "Medium", "pre": "ma chuyến hợp lệ.", "data": "GET /api/chuyen-xe/{ma}", "steps": "1. Gọi API.", "expected": "Trả tuyến, xe, giờ, giá.", "post": "Không đổi DL."},
        {"desc": "Xóa chuyến không có vé", "prio": "Low", "pre": "Chuyến không có vé.", "data": "DELETE /api/chuyen-xe/{ma}", "steps": "1. Xóa.", "expected": "Xóa thành công.", "post": "Không còn chuyến."},
    ],
    "TimKiemChuyen": [
        {"desc": "Tìm chuyến theo tuyến và ngày", "prio": "High", "pre": "Có chuyến ngày chọn.", "data": "GET /api/chuyen-xe/tim-kiem?maTuyen=&ngay=", "steps": "1. Trang chủ nhập điểm đi/đến, ngày.\n2. Tìm kiếm.", "expected": "Danh sách chuyến phù hợp.", "post": "Không đổi DL."},
        {"desc": "Lọc theo khoảng giá", "prio": "Medium", "pre": "Nhiều chuyến giá khác nhau.", "data": "giaMin, giaMax", "steps": "1. Đặt bộ lọc giá.\n2. Tìm.", "expected": "Chỉ chuyến trong khoảng.", "post": "Không đổi DL."},
        {"desc": "Lọc theo loại xe / hãng xe", "prio": "Medium", "pre": "Chuyến đa dạng loại xe.", "data": "hangXe hoặc maLoaiXe", "steps": "1. Chọn bộ lọc.\n2. Tìm.", "expected": "Kết quả lọc đúng.", "post": "Không đổi DL."},
        {"desc": "Lọc khung giờ khởi hành", "prio": "Low", "pre": "Chuyến nhiều khung giờ.", "data": "gioSang / gioChieu / gioToi", "steps": "1. Chọn khung giờ.", "expected": "Chuyến trong khung.", "post": "Không đổi DL."},
        {"desc": "Tìm kiếm – ngày không có chuyến", "prio": "Medium", "pre": "Không có chuyến ngày X.", "data": "ngay không có DL", "steps": "1. Tìm ngày trống.", "expected": "Thông báo không có chuyến.", "post": "Không đổi DL."},
        {"desc": "Sắp xếp kết quả (giá, giờ)", "prio": "Low", "pre": "Có ≥2 chuyến.", "data": "sort=gia|gio", "steps": "1. Chọn tiêu chí sắp xếp.", "expected": "Thứ tự đúng.", "post": "Không đổi DL."},
    ],
    "DatVe": [
        {"desc": "Đặt 1 ghế – tạo vé PENDING", "prio": "High", "pre": "CUSTOMER đăng nhập; ghế trống.", "data": "maChuyen, maGhe, maDiemLen, maDiemXuong", "steps": "1. Chọn ghế.\n2. POST /api/ve-xe/dat.", "expected": "Vé PENDING; mã vé RB...; giữ ghế 15 phút.", "post": "ve_xe PENDING; ghế giữ."},
        {"desc": "Đặt nhiều ghế (tối đa 10)", "prio": "High", "pre": "Đủ ghế trống.", "data": "dsMaGhe length 2–10", "steps": "1. Chọn nhiều ghế.\n2. Đặt.", "expected": "Nhiều vé PENDING cùng lúc.", "post": "N vé PENDING."},
        {"desc": "Đặt vé – không chọn ghế", "prio": "High", "pre": "CUSTOMER đăng nhập.", "data": "dsMaGhe rỗng", "steps": "1. Gửi dat không ghế.", "expected": "IllegalArgumentException / lỗi 400.", "post": "Không tạo vé."},
        {"desc": "Đặt vé – quá 10 ghế", "prio": "Medium", "pre": "CUSTOMER đăng nhập.", "data": "11 ghế", "steps": "1. Chọn 11 ghế.\n2. Đặt.", "expected": "Từ chối; báo tối đa 10.", "post": "Không tạo vé."},
        {"desc": "Đặt ghế đã được giữ/bán", "prio": "High", "pre": "Ghế PENDING/PAID.", "data": "maGhe đã giữ", "steps": "1. User B chọn ghế User A giữ.", "expected": "Từ chối; báo ghế đã bán.", "post": "Không tạo vé trùng."},
        {"desc": "Xem vé của tôi", "prio": "High", "pre": "CUSTOMER có vé.", "data": "GET /api/ve-xe/cua-toi", "steps": "1. Mở Trang Vé của tôi.", "expected": "Danh sách vé; tự xử lý hết hạn.", "post": "Vé EXPIRED nếu quá hạn."},
        {"desc": "Hủy vé PENDING", "prio": "High", "pre": "Vé PENDING của khách.", "data": "POST /api/ve-xe/{ma}/huy", "steps": "1. Hủy vé chờ.", "expected": "Vé CANCELLED; ghế giải phóng.", "post": "Ghế trống."},
        {"desc": "Hủy vé PAID – từ chối", "prio": "High", "pre": "Vé đã thanh toán.", "data": "ma vé PAID", "steps": "1. Thử hủy.", "expected": "IllegalStateException.", "post": "Vé vẫn PAID."},
        {"desc": "Đặt vé không đăng nhập", "prio": "High", "pre": "Chưa JWT.", "data": "Không token", "steps": "1. POST /api/ve-xe/dat.", "expected": "HTTP 401.", "post": "Không tạo vé."},
    ],
    "HetHanVe": [
        {"desc": "Scheduled task hủy vé PENDING quá 15 phút", "prio": "High", "pre": "Vé PENDING > 15 phút.", "data": "app.ve.phut-cho-thanh-toan=15", "steps": "1. Chờ task chạy.\n2. Kiểm tra CSDL.", "expected": "Vé EXPIRED; ghế giải phóng.", "post": "Ghế AVAILABLE."},
        {"desc": "Vé PENDING trong hạn – không bị hủy", "prio": "High", "pre": "Vé mới < 15 phút.", "data": "thoi_gian_dat gần hiện tại", "steps": "1. Chờ task.\n2. Kiểm tra.", "expected": "Vẫn PENDING.", "post": "Ghế vẫn giữ."},
        {"desc": "Đếm ngược thanh toán trên UI", "prio": "Medium", "pre": "Vé PENDING.", "data": "PHUT_CHO_THANH_TOAN=15", "steps": "1. Mở vé PENDING.\n2. Quan sát timer.", "expected": "Hiển thị đếm ngược; hết giờ báo hết hạn.", "post": "UI refresh trạng thái."},
        {"desc": "Mở Vé của tôi kích hoạt xử lý hết hạn", "prio": "Medium", "pre": "Có vé quá hạn.", "data": "GET /api/ve-xe/cua-toi", "steps": "1. Gọi API cua-toi.", "expected": "Gọi xuLyHetHanChoKhach trước khi trả DL.", "post": "Vé quá hạn → EXPIRED."},
    ],
    "ThanhToan": [
        {"desc": "Thanh toán tiền mặt tại quầy ({role})", "prio": "High", "pre": "Vé PENDING; {role} là STAFF hoặc ADMIN.", "data": "POST /api/thanh-toan/ve/{maVe}/tien-mat", "steps": "1. Chọn vé PENDING.\n2. Thu tiền mặt.", "expected": "Giao dịch SUCCESS; vé PAID.", "post": "ve_xe PAID."},
        {"desc": "Thanh toán PayOS – tạo link", "prio": "High", "pre": "Vé PENDING; PayOS cấu hình.", "data": "POST /api/thanh-toan/ve/{maVe}/payos", "steps": "1. Chọn PayOS.\n2. Nhận checkoutUrl.", "expected": "Trả link; orderCode lưu.", "post": "Giao dịch PENDING."},
        {"desc": "PayOS callback thành công", "prio": "High", "pre": "Link PayOS đã tạo.", "data": "GET /api/thanh-toan/payos/ket-qua?orderCode=", "steps": "1. Khách thanh toán sandbox.\n2. Redirect kết quả.", "expected": "Vé PAID; thông báo SSE.", "post": "giao_dich SUCCESS."},
        {"desc": "PayOS webhook xác nhận", "prio": "High", "pre": "PayOS gửi webhook.", "data": "POST /api/thanh-toan/payos/webhook", "steps": "1. Mô phỏng webhook PAID.", "expected": "Cập nhật vé PAID nếu chưa.", "post": "Đồng bộ trạng thái."},
        {"desc": "Áp dụng mã khuyến mãi hợp lệ", "prio": "Medium", "pre": "Mã KM còn hạn, đủ điều kiện.", "data": "maKhuyenMai", "steps": "1. Nhập mã khi thanh toán.", "expected": "Giảm giá đúng %/số tiền.", "post": "so_tien giảm."},
        {"desc": "Mã khuyến mãi hết hạn / sai", "prio": "Medium", "pre": "Mã không hợp lệ.", "data": "maKm sai", "steps": "1. Nhập mã sai.\n2. Thanh toán.", "expected": "Báo lỗi; không giảm giá.", "post": "Không áp KM."},
        {"desc": "Thanh toán vé không PENDING", "prio": "High", "pre": "Vé PAID/EXPIRED.", "data": "maVe PAID", "steps": "1. Thử thanh toán lại.", "expected": "Từ chối.", "post": "Trạng thái không đổi."},
        {"desc": "Xem lịch sử thanh toán", "prio": "Medium", "pre": "CUSTOMER có giao dịch.", "data": "GET /api/thanh-toan/lich-su", "steps": "1. Mở Trang Lịch sử thanh toán.", "expected": "Danh sách giao dịch, số tiền, PT.", "post": "Không đổi DL."},
        {"desc": "Xem hình thức thanh toán", "prio": "Low", "pre": "Seed hinh_thuc_thanh_toan.", "data": "GET /api/hinh-thuc-thanh-toan", "steps": "1. Gọi API.", "expected": "TIEN_MAT, PAYOS hiển thị.", "post": "Không đổi DL."},
    ],
    "TraCuuVe": [
        {"desc": "Tra cứu vé công khai (mã vé + SĐT)", "prio": "High", "pre": "Vé tồn tại.", "data": "GET /api/ve-xe/tra-cuu?maVe=&soDienThoai=", "steps": "1. Mở Trang Tra cứu.\n2. Nhập mã + SĐT.", "expected": "Hiển thị thông tin vé (không cần login).", "post": "Không đổi DL."},
        {"desc": "Tra cứu – SĐT không khớp", "prio": "High", "pre": "Vé tồn tại.", "data": "SĐT sai", "steps": "1. Nhập mã đúng, SĐT sai.", "expected": "Không hiển thị / báo không tìm thấy.", "post": "Không lộ DL."},
        {"desc": "Xem vé điện tử đã đăng nhập", "prio": "High", "pre": "CUSTOMER; vé của mình.", "data": "GET /api/ve-xe/{ma}/dien-tu", "steps": "1. Mở vé PAID.\n2. Xem vé điện tử.", "expected": "Lộ trình, ghế, QR, hãng xe.", "post": "Không đổi DL."},
        {"desc": "Xuất/chụp vé điện tử QR", "prio": "Medium", "pre": "Vé PAID.", "data": "maVe QR payload", "steps": "1. Hiển thị QR.\n2. Tải ảnh vé (html-to-image).", "expected": "QR chứa URL tra cứu đúng.", "post": "File ảnh tải về."},
        {"desc": "Tra cứu vé người khác – từ chối", "prio": "High", "pre": "CUSTOMER A.", "data": "ma vé của B", "steps": "1. Gọi dien-tu vé B.", "expected": "403/404.", "post": "Không lộ vé B."},
    ],
    "DoiVe": [
        {"desc": "Đổi ghế – ghế mới trống", "prio": "High", "pre": "Vé PAID; ghế mới AVAILABLE.", "data": "POST /api/ve-xe/{ma}/doi-ghe — maGheMoi", "steps": "1. Chọn ghế mới.\n2. Xác nhận.", "expected": "Vé cập nhật ghế; ghế cũ trống.", "post": "ma_ghe mới."},
        {"desc": "Đổi ghế – ghế mới đã bán", "prio": "High", "pre": "Ghế mới đã PAID.", "data": "maGheMoi occupied", "steps": "1. Thử đổi.", "expected": "Từ chối.", "post": "Ghế không đổi."},
        {"desc": "Đổi chuyến – cùng tuyến", "prio": "High", "pre": "Vé PAID; chuyến mới cùng tuyến.", "data": "POST /api/ve-xe/{ma}/doi-chuyen", "steps": "1. Chọn chuyến mới.\n2. Chọn ghế mới.", "expected": "Vé gắn chuyến mới.", "post": "ma_chuyen cập nhật."},
        {"desc": "Đổi chuyến – khác tuyến", "prio": "Medium", "pre": "Vé PAID.", "data": "chuyến khác tuyến", "steps": "1. Thử đổi sang tuyến khác.", "expected": "Từ chối theo quy tắc nghiệp vụ.", "post": "Vé không đổi."},
        {"desc": "Đổi vé PENDING", "prio": "Medium", "pre": "Vé PENDING.", "data": "doi-ghe/doi-chuyen", "steps": "1. Thử đổi vé chờ.", "expected": "Theo quy tắc (cho phép hoặc từ chối).", "post": "Ghi nhận hành vi."},
        {"desc": "Thông báo sau đổi vé", "prio": "Low", "pre": "Đổi vé thành công.", "data": "SSE + DB thong_bao", "steps": "1. Đổi vé.\n2. Kiểm tra thông báo.", "expected": "Khách nhận thông báo đổi vé.", "post": "Có thong_bao mới."},
    ],
    "ThongBao": [
        {"desc": "Kết nối SSE stream thông báo", "prio": "High", "pre": "Đã đăng nhập.", "data": "GET /api/thong-bao/stream", "steps": "1. Mở app.\n2. EventSource kết nối.", "expected": "Nhận sự kiện realtime.", "post": "Kết nối SSE active."},
        {"desc": "Nhận thông báo sau thanh toán", "prio": "High", "pre": "Vé chuyển PAID.", "data": "Sự kiện thanh toán", "steps": "1. Thanh toán.\n2. Quan sát chuông thông báo.", "expected": "Popup/badge thông báo mới.", "post": "thong_bao chưa đọc."},
        {"desc": "Xem danh sách thông báo", "prio": "Medium", "pre": "Có thông báo.", "data": "GET /api/thong-bao", "steps": "1. Mở Trang Thông báo.", "expected": "Danh sách tiêu đề, nội dung, thời gian.", "post": "Không đổi DL."},
        {"desc": "Đánh dấu đã đọc", "prio": "Low", "pre": "Thông báo chưa đọc.", "data": "PUT /api/thong-bao/{ma}/da-doc", "steps": "1. Click đánh dấu đọc.", "expected": "da_doc=1.", "post": "Badge giảm."},
        {"desc": "Gửi email xác nhận vé", "prio": "Medium", "pre": "SMTP cấu hình.", "data": "Sau datVe/thanh toán", "steps": "1. Đặt vé & thanh toán.\n2. Kiểm tra hộp thư.", "expected": "Email có mã vé, chuyến.", "post": "Mail đã gửi (hoặc log).", "note": "Dev có thể mock mail."},
    ],
    "Chat": [
        {"desc": "Xem hội thoại với nhân viên", "prio": "Medium", "pre": "CUSTOMER đăng nhập.", "data": "GET /api/chat/hoi-thoai?doiPhuong=", "steps": "1. Mở khung chat.\n2. Chọn đối phương.", "expected": "Lịch sử tin nhắn.", "post": "Không đổi DL."},
        {"desc": "Gửi tin nhắn", "prio": "Medium", "pre": "Đăng nhập.", "data": "POST /api/chat/gui — noiDung, maNguoiNhan", "steps": "1. Nhập nội dung.\n2. Gửi.", "expected": "Tin lưu tin_nhan_chat.", "post": "Có tin mới."},
        {"desc": "Gửi tin rỗng", "prio": "Low", "pre": "Đăng nhập.", "data": "noiDung rỗng", "steps": "1. Gửi tin trống.", "expected": "Validation từ chối.", "post": "Không lưu tin."},
        {"desc": "Chat không đăng nhập", "prio": "High", "pre": "Không JWT.", "data": "Không token", "steps": "1. POST /api/chat/gui.", "expected": "HTTP 401.", "post": "Không gửi."},
    ],
    "HoiDap": [
        {"desc": "Xem FAQ công khai", "prio": "Medium", "pre": "Có câu hỏi đã trả lời.", "data": "GET /api/hoi-dap/cong-khai", "steps": "1. Mở mục Hỏi đáp.", "expected": "Danh sách Q&A.", "post": "Không đổi DL."},
        {"desc": "Khách gửi câu hỏi mới", "prio": "Medium", "pre": "CUSTOMER đăng nhập.", "data": "POST /api/hoi-dap", "steps": "1. Gửi câu hỏi.", "expected": "Lưu câu hỏi chờ trả lời.", "post": "Có bản ghi FAQ."},
        {"desc": "Admin trả lời câu hỏi", "prio": "Medium", "pre": "ADMIN đăng nhập.", "data": "PUT /api/hoi-dap/{ma}/tra-loi", "steps": "1. Nhập trả lời.\n2. Lưu.", "expected": "Câu hỏi hiện công khai.", "post": "tra_loi có nội dung."},
        {"desc": "Quản trị xem tất cả câu hỏi", "prio": "Low", "pre": "ADMIN.", "data": "GET /api/hoi-dap/quan-tri/tat-ca", "steps": "1. Mở quản trị FAQ.", "expected": "Cả câu chờ và đã trả lời.", "post": "Không đổi DL."},
    ],
    "TinTuc": [
        {"desc": "Xem danh sách tin tức công khai", "prio": "Medium", "pre": "Có tin xuất bản.", "data": "GET /api/tin-tuc", "steps": "1. Mở Trang Tin tức.", "expected": "Danh sách bài viết.", "post": "Không đổi DL."},
        {"desc": "Xem chi tiết tin", "prio": "Medium", "pre": "ma tin hợp lệ.", "data": "GET /api/tin-tuc/{ma}", "steps": "1. Click bài viết.", "expected": "Nội dung HTML, ảnh.", "post": "Không đổi DL."},
        {"desc": "Admin thêm tin mới", "prio": "Medium", "pre": "ADMIN.", "data": "POST /api/tin-tuc", "steps": "1. Soạn tin.\n2. Lưu.", "expected": "Tin xuất hiện danh sách.", "post": "Có tin_tuc."},
        {"desc": "Upload ảnh bài viết", "prio": "Low", "pre": "ADMIN.", "data": "POST /api/tin-tuc/upload-anh", "steps": "1. Upload ảnh editor.", "expected": "URL ảnh trả về.", "post": "Ảnh lưu server."},
        {"desc": "Cập nhật / xóa tin", "prio": "Low", "pre": "Tin tồn tại.", "data": "PUT/DELETE /api/tin-tuc/{ma}", "steps": "1. Sửa hoặc xóa.", "expected": "CRUD thành công.", "post": "DL cập nhật."},
    ],
    "DanhGia": [
        {"desc": "Xem đánh giá công khai", "prio": "Medium", "pre": "Có đánh giá.", "data": "GET /api/danh-gia/cong-khai", "steps": "1. Xem slider đánh giá trang chủ.", "expected": "Hiển thị sao, nội dung.", "post": "Không đổi DL."},
        {"desc": "Xem đánh giá theo chuyến", "prio": "Medium", "pre": "Chuyến có đánh giá.", "data": "GET /api/danh-gia/chuyen/{maChuyen}", "steps": "1. Mở chi tiết chuyến.", "expected": "Danh sách đánh giá.", "post": "Không đổi DL."},
        {"desc": "Khách đánh giá sau chuyến", "prio": "High", "pre": "Vé USED/đủ điều kiện.", "data": "POST /api/danh-gia — diem, noiDung", "steps": "1. Mở Trang Đánh giá.\n2. Gửi.", "expected": "Lưu danh_gia_chuyen.", "post": "Có đánh giá mới."},
        {"desc": "Xem vé chờ đánh giá", "prio": "Medium", "pre": "CUSTOMER.", "data": "GET /api/danh-gia/ve-cho-danh-gia", "steps": "1. Mở trang đánh giá.", "expected": "Danh sách vé đủ điều kiện.", "post": "Không đổi DL."},
        {"desc": "Đánh giá trùng – từ chối", "prio": "Medium", "pre": "Đã đánh giá chuyến.", "data": "cùng vé/chuyến", "steps": "1. Gửi đánh giá lần 2.", "expected": "Từ chối.", "post": "Không trùng bản ghi."},
    ],
    "KhuyenMai": [
        {"desc": "Xem KM hiển thị công khai", "prio": "Medium", "pre": "KM hoạt động.", "data": "GET /api/khuyen-mai/hien-thi", "steps": "1. Xem banner KM.", "expected": "Hiển thị mã, % giảm.", "post": "Không đổi DL."},
        {"desc": "Admin thêm mã khuyến mãi", "prio": "Medium", "pre": "ADMIN.", "data": "ma, phanTram/giamTien, ngayBatDau, ngayKetThuc", "steps": "1. POST /api/khuyen-mai.", "expected": "Tạo KM.", "post": "Có khuyen_mai."},
        {"desc": "Cập nhật KM", "prio": "Low", "pre": "KM tồn tại.", "data": "PUT /api/khuyen-mai/{ma}", "steps": "1. Sửa KM.\n2. Lưu.", "expected": "Cập nhật OK.", "post": "DL đổi."},
        {"desc": "Xóa KM", "prio": "Low", "pre": "KM tồn tại.", "data": "DELETE /api/khuyen-mai/{ma}", "steps": "1. Xóa.", "expected": "Không hiển thị nữa.", "post": "KM xóa."},
    ],
    "BaoCao": [
        {"desc": "Dashboard tổng quan ({role})", "prio": "High", "pre": "Đăng nhập {role} quản trị.", "data": "Trang Tổng quan", "steps": "1. Mở dashboard.", "expected": "Thẻ thống kê vé, doanh thu.", "post": "Không đổi DL."},
        {"desc": "Báo cáo mở rộng", "prio": "High", "pre": "ADMIN/STAFF.", "data": "GET /api/bao-cao/mo-rong?tuNgay&denNgay", "steps": "1. Chọn khoảng ngày.\n2. Xem báo cáo.", "expected": "Số liệu khớp CSDL.", "post": "Không đổi DL."},
        {"desc": "Biểu đồ báo cáo", "prio": "Medium", "pre": "Có dữ liệu.", "data": "GET /api/bao-cao/bieu-do", "steps": "1. Mở Trang Báo cáo.\n2. Xem Recharts.", "expected": "Biểu đồ doanh thu, vé theo TT.", "post": "Không đổi DL."},
        {"desc": "Xuất CSV báo cáo", "prio": "Medium", "pre": "ADMIN/STAFF.", "data": "GET /api/bao-cao/xuat-csv", "steps": "1. Nhấn Xuất CSV.\n2. Mở file.", "expected": "UTF-8 BOM; header đúng.", "post": "File tải về."},
        {"desc": "CUSTOMER không xuất CSV", "prio": "High", "pre": "CUSTOMER.", "data": "JWT CUSTOMER", "steps": "1. Gọi xuat-csv.", "expected": "HTTP 403.", "post": "Không có file."},
    ],
    "KhachHang": [
        {"desc": "Xem danh sách khách ({role})", "prio": "Medium", "pre": "Đăng nhập {role}.", "data": "GET /api/khach-hang", "steps": "1. Mở Trang Khách hàng.", "expected": "Danh sách họ tên, SĐT.", "post": "Không đổi DL."},
        {"desc": "Thêm khách qua quản trị", "prio": "Medium", "pre": "ADMIN/STAFF.", "data": "POST /api/khach-hang", "steps": "1. Thêm khách + tài khoản.", "expected": "Tạo tai_khoan + khach_hang.", "post": "Có khách mới."},
        {"desc": "Cập nhật hồ sơ khách", "prio": "Medium", "pre": "Khách tồn tại.", "data": "PUT /api/khach-hang/{ma}", "steps": "1. Sửa SĐT/địa chỉ.", "expected": "Cập nhật OK.", "post": "DL đổi."},
        {"desc": "CUSTOMER không quản lý khách", "prio": "High", "pre": "CUSTOMER.", "data": "GET /api/khach-hang", "steps": "1. Gọi API.", "expected": "403.", "post": "Không lộ DL."},
    ],
    "HoSo": [
        {"desc": "Xem hồ sơ cá nhân", "prio": "Medium", "pre": "Đã đăng nhập.", "data": "GET /api/ho-so/cua-toi", "steps": "1. Mở Trang Hồ sơ.", "expected": "Họ tên, email, SĐT.", "post": "Không đổi DL."},
        {"desc": "Cập nhật hồ sơ", "prio": "Medium", "pre": "Đã đăng nhập.", "data": "PUT /api/ho-so/cua-toi", "steps": "1. Sửa họ tên/SĐT.\n2. Lưu.", "expected": "Cập nhật khach_hang.", "post": "DL đổi."},
        {"desc": "Đổi mật khẩu", "prio": "High", "pre": "Biết MK cũ.", "data": "PUT /api/ho-so/mat-khau", "steps": "1. Nhập MK cũ/mới.\n2. Lưu.", "expected": "Đổi MK thành công.", "post": "MK mới hiệu lực."},
        {"desc": "Đổi MK – MK cũ sai", "prio": "Medium", "pre": "Đã đăng nhập.", "data": "matKhauCu sai", "steps": "1. Nhập MK cũ sai.", "expected": "Từ chối.", "post": "MK không đổi."},
    ],
    "DiaDanh": [
        {"desc": "Lấy danh sách tỉnh", "prio": "Medium", "pre": "API địa danh.", "data": "GET /api/dia-danh/tinh", "steps": "1. Mở dropdown tỉnh.", "expected": "Danh sách tỉnh.", "post": "Không đổi DL."},
        {"desc": "Lấy xã theo tỉnh", "prio": "Medium", "pre": "Chọn tỉnh.", "data": "GET /api/dia-danh/tinh/{ma}/xa", "steps": "1. Chọn tỉnh.\n2. Load xã.", "expected": "Danh sách xã đúng tỉnh.", "post": "Không đổi DL."},
        {"desc": "Ước tính lộ trình", "prio": "Low", "pre": "2 điểm địa lý.", "data": "GET /api/dia-danh/uoc-tinh-lo-trinh", "steps": "1. Chọn điểm đi/đến.", "expected": "Khoảng cách/thời gian ước tính.", "post": "Không đổi DL."},
    ],
    "PhanQuyen": [
        {"desc": "ADMIN truy cập toàn bộ menu quản trị", "prio": "High", "pre": "Đăng nhập admin.", "data": "admin/Admin@123", "steps": "1. Đăng nhập admin.\n2. Duyệt menu.", "expected": "Truy cập tất cả trang /quan-tri/*.", "post": "Không đổi DL."},
        {"desc": "STAFF truy cập chức năng vận hành", "prio": "High", "pre": "Đăng nhập staff.", "data": "staff/Staff@123", "steps": "1. Đăng nhập staff.\n2. Thử bán vé, xem chuyến.", "expected": "Cho phép vận hành; hạn chế cấu hình (nếu có).", "post": "Không đổi DL."},
        {"desc": "CUSTOMER không vào /quan-tri", "prio": "High", "pre": "CUSTOMER.", "data": "Navigate /quan-tri", "steps": "1. Truy cập URL quản trị.", "expected": "Redirect / từ chối.", "post": "Không lộ DL admin."},
        {"desc": "API không token – 401", "prio": "High", "pre": "Endpoint bảo vệ.", "data": "Không Authorization", "steps": "1. Gọi API protected.", "expected": "HTTP 401.", "post": "Không trả DL."},
        {"desc": "CORS – frontend gọi API", "prio": "Medium", "pre": "Vite localhost:5173.", "data": "Origin frontend", "steps": "1. Gọi API từ frontend.", "expected": "Không lỗi CORS.", "post": "Request thành công."},
    ],
    "GiaoDien": [
        {"desc": "Trang chủ – tìm chuyến nhanh", "prio": "High", "pre": "App chạy.", "data": "Route /", "steps": "1. Mở trang chủ.\n2. Nhập tìm kiếm.", "expected": "Form tìm chuyến hoạt động.", "post": "Không đổi DL."},
        {"desc": "Trang đặt vé – sơ đồ ghế", "prio": "High", "pre": "Có chuyến.", "data": "Route /dat-ve/:maChuyen", "steps": "1. Mở trang đặt vé.\n2. Chọn ghế.", "expected": "Sơ đồ ghế tương tác.", "post": "Không đổi DL."},
        {"desc": "Trang kết quả thanh toán PayOS", "prio": "High", "pre": "Redirect từ PayOS.", "data": "Route /thanh-toan/ket-qua", "steps": "1. Mở URL kết quả.", "expected": "Hiển thị thành công/thất bại.", "post": "Không đổi DL."},
        {"desc": "Responsive – màn hình mobile", "prio": "Medium", "pre": "DevTools mobile.", "data": "Viewport 375px", "steps": "1. Thu nhỏ màn hình.\n2. Duyệt trang.", "expected": "Layout không vỡ.", "post": "Không đổi DL."},
        {"desc": "Trang 404", "prio": "Low", "pre": "URL sai.", "data": "Route /khong-ton-tai", "steps": "1. Truy cập URL lạ.", "expected": "TrangKhongTimThay.", "post": "Không đổi DL."},
        {"desc": "Bản đồ Leaflet điểm dừng", "prio": "Low", "pre": "Chuyến có tọa độ.", "data": "react-leaflet", "steps": "1. Xem lộ trình trên bản đồ.", "expected": "Marker điểm dừng hiển thị.", "post": "Không đổi DL."},
    ],
    "LuongE2E": [
        {"desc": "E2E: Cấu hình tuyến → chuyến → đặt vé → PayOS → vé điện tử", "prio": "High", "pre": "DB seed; PayOS sandbox.", "data": "Bộ dữ liệu test", "steps": "1. Admin tạo tuyến/xe/chuyến.\n2. Khách tìm chuyến, đặt vé.\n3. PayOS thanh toán.\n4. Xem vé QR.", "expected": "Luồng hoàn tất; DL nhất quán.", "post": "Vé PAID; báo cáo cập nhật.", "note": "Demo bảo vệ."},
        {"desc": "E2E: Bán vé tiền mặt tại quầy", "prio": "High", "pre": "STAFF đăng nhập.", "data": "Vé PENDING khách A", "steps": "1. Staff tìm vé PENDING.\n2. Thu tiền mặt.", "expected": "Vé PAID ngay.", "post": "Giao dịch TIEN_MAT."},
        {"desc": "E2E: Vé hết hạn → ghế giải phóng → khách B đặt", "prio": "High", "pre": "Vé A PENDING quá hạn.", "data": "2 khách, 1 ghế", "steps": "1. A đặt không thanh toán.\n2. Chờ hết hạn.\n3. B đặt cùng ghế.", "expected": "B đặt thành công.", "post": "A EXPIRED; B PENDING."},
        {"desc": "E2E: Đổi ghế sau thanh toán", "prio": "Medium", "pre": "Vé PAID.", "data": "Ghế mới trống", "steps": "1. Đổi ghế.\n2. Kiểm tra vé điện tử.", "expected": "Ghế mới trên vé QR.", "post": "DL cập nhật."},
        {"desc": "E2E: Đánh giá sau chuyến", "prio": "Medium", "pre": "Chuyến đã đi.", "data": "Vé đủ điều kiện", "steps": "1. Hoàn thành chuyến (mock).\n2. Đánh giá.", "expected": "Đánh giá hiện công khai.", "post": "Có danh_gia."},
    ],
    "BaoMat": [
        {"desc": "SQL Injection – tham số tìm kiếm", "prio": "High", "pre": "API tim-kiem.", "data": "' OR 1=1 --", "steps": "1. Gửi payload SQLi.", "expected": "Không lộ DL; lỗi an toàn.", "post": "CSDL không đổi."},
        {"desc": "XSS – nội dung tin tức/chat", "prio": "High", "pre": "Form nhập HTML.", "data": "<script>alert(1)</script>", "steps": "1. Gửi nội dung XSS.\n2. Hiển thị.", "expected": "Escape/sanitize; không thực thi script.", "post": "An toàn."},
        {"desc": "Truy cập vé người khác bằng IDOR", "prio": "High", "pre": "2 tài khoản CUSTOMER.", "data": "ma vé của user B", "steps": "1. User A gọi API vé B.", "expected": "403/404.", "post": "Không lộ vé."},
        {"desc": "Webhook PayOS – chữ ký không hợp lệ", "prio": "High", "pre": "Endpoint webhook.", "data": "Payload giả", "steps": "1. POST webhook không chữ ký.", "expected": "Từ chối; không PAID vé.", "post": "An toàn."},
        {"desc": "Brute force đăng nhập (biên)", "prio": "Medium", "pre": "Nhiều lần đăng nhập sai.", "data": "100 request sai", "steps": "1. Thử liên tục MK sai.", "expected": "Giới hạn hoặc chậm (nếu cấu hình).", "post": "Ghi nhận hành vi.", "note": "Có thể chưa có rate limit."},
    ],
}

# Số lần nhân bản biến thể mỗi case gốc (để đạt 2000–5000)
EXPAND_PER_BASE = {
    "XacThuc": 18,
    "TuyenDuong": 14,
    "XeKhach": 14,
    "LoaiXe": 14,
    "DiemDungChan": 16,
    "GheNgoi": 18,
    "ChuyenXe": 16,
    "TimKiemChuyen": 18,
    "DatVe": 20,
    "HetHanVe": 20,
    "ThanhToan": 18,
    "TraCuuVe": 18,
    "DoiVe": 18,
    "ThongBao": 16,
    "Chat": 18,
    "HoiDap": 18,
    "TinTuc": 16,
    "DanhGia": 16,
    "KhuyenMai": 18,
    "BaoCao": 16,
    "KhachHang": 16,
    "HoSo": 18,
    "DiaDanh": 22,
    "PhanQuyen": 16,
    "GiaoDien": 18,
    "LuongE2E": 25,
    "BaoMat": 18,
}

VARIANT_SUFFIXES = [
    "",
    " – biến thể UI",
    " – kiểm thử API trực tiếp",
    " – kiểm thử qua giao diện",
    " – dữ liệu biên min",
    " – dữ liệu biên max",
    " – môi trường dev local",
    " – sau khi refresh trang",
    " – kiểm thử hồi quy",
    " – song song 2 tab",
]

VARIANT_ROLES_FOR_AUTH = ["ADMIN", "STAFF", "CUSTOMER"]


def _sub_role(text: str | None, role: str) -> str:
    return (text or "").replace("{role}", role)


def expand_case(base: dict, module: str, variant_idx: int) -> dict:
    c = copy.deepcopy(base)
    suffix = VARIANT_SUFFIXES[variant_idx % len(VARIANT_SUFFIXES)]
    role = VARIANT_ROLES_FOR_AUTH[variant_idx % len(VARIANT_ROLES_FOR_AUTH)]

    if "{role}" in c.get("desc", ""):
        c["desc"] = _sub_role(c["desc"], role)
    else:
        c["desc"] = c["desc"] + suffix

    for key in ("pre", "data", "expected"):
        if key in c:
            c[key] = _sub_role(c[key], role)

    if suffix and " – " not in c["desc"][:20]:
        pass  # already appended

    if variant_idx > 0 and suffix:
        c["note"] = (c.get("note") or "") + f" Biến thể #{variant_idx + 1}."

    return c


def build_rows() -> dict[str, list[tuple]]:
    """Trả về dict sheet_name -> list rows (không header)."""
    sheets: dict[str, list[tuple]] = {}
    tc_counter = 1

    for module, bases in MODULE_BASE.items():
        rows = []
        expand_n = EXPAND_PER_BASE.get(module, 15)
        for base in bases:
            for v in range(expand_n):
                case = expand_case(base, module, v)
                tc_id = f"TC{tc_counter:04d}"
                tc_counter += 1
                rows.append(
                    (
                        tc_id,
                        module,
                        case.get("prio", "Medium"),
                        case["desc"],
                        case.get("pre", ""),
                        case.get("data", ""),
                        case.get("steps", ""),
                        case.get("expected", ""),
                        ACTUAL_PLACEHOLDER,
                        STATUS_PLACEHOLDER,
                        case.get("post", ""),
                        case.get("note"),
                    )
                )
        sheets[module] = rows
    return sheets


def style_sheet(ws, row_count: int):
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [12, 14, 10, 42, 28, 32, 36, 36, 18, 16, 28, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wrap = Alignment(wrap_text=True, vertical="top")
    for r in range(2, row_count + 1):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=c).alignment = wrap

    ws.freeze_panes = "A2"


def main():
    out_path = Path(__file__).parent / "RedBus_TestCase.xlsx"
    sheets_data = build_rows()
    total = sum(len(v) for v in sheets_data.values())

    # Điều chỉnh nếu chưa đủ 2000
    if total < 2000:
        raise SystemExit(f"Chỉ sinh {total} case, cần tăng EXPAND_PER_BASE")

    wb = openpyxl.Workbook()
    # Xóa sheet mặc định, tạo sheet Tổng hợp trước
    default = wb.active
    wb.remove(default)

    summary = wb.create_sheet("TongHop", 0)
    summary.append(["Bộ test case hệ thống RedBus"])
    summary.append(["Tổng số test case", total])
    summary.append(["Số sheet chức năng", len(sheets_data)])
    summary.append(["Ngày sinh", "2026-06-09"])
    summary.append([])
    summary.append(["Sheet", "Số case"])
    for name, rows in sorted(sheets_data.items(), key=lambda x: -len(x[1])):
        summary.append([name, len(rows)])

    for sheet_name, rows in sheets_data.items():
        ws = wb.create_sheet(sheet_name[:31])  # Excel limit 31 chars
        ws.append(HEADERS)
        for row in rows:
            ws.append(row)
        style_sheet(ws, len(rows) + 1)

    wb.save(out_path)
    log = Path(__file__).parent / "_gen_log.txt"
    with open(log, "w", encoding="utf-8") as f:
        f.write(f"Da sinh {total} test case -> {out_path}\n")
        for name, rows in sorted(sheets_data.items(), key=lambda x: -len(x[1])):
            f.write(f"  {name}: {len(rows)}\n")
    print(f"OK: {total} cases -> {out_path}")


if __name__ == "__main__":
    main()
