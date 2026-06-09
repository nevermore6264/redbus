# -*- coding: utf-8 -*-
import openpyxl
wb = openpyxl.load_workbook(r'd:\Freelancer\RedBus\test-case\TestCase.xlsx')
ws = wb['TestCase']
with open(r'd:\Freelancer\RedBus\test-case\_inspect.txt', 'w', encoding='utf-8') as f:
    f.write(f'rows: {ws.max_row}\n')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        f.write(f'{i}: {row}\n')
    f.write('\n--- samples from middle ---\n')
    for i, row in enumerate(ws.iter_rows(min_row=500, max_row=505, values_only=True), 500):
        f.write(f'{i}: {row}\n')
    f.write('\n--- last rows ---\n')
    for i, row in enumerate(ws.iter_rows(min_row=ws.max_row-3, max_row=ws.max_row, values_only=True), ws.max_row-3):
        f.write(f'{i}: {row}\n')
