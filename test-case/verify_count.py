# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path
p = Path(__file__).parent / "RedBus_TestCase.xlsx"
wb = openpyxl.load_workbook(p, read_only=True)
total = 0
lines = []
for name in wb.sheetnames:
    ws = wb[name]
    n = ws.max_row - 1 if ws.max_row else 0
    if name != "TongHop":
        total += n
    lines.append(f"{name}: {n}")
with open(Path(__file__).parent / "_verify.txt", "w", encoding="utf-8") as f:
    f.write(f"TOTAL: {total}\n")
    f.write("\n".join(lines))
